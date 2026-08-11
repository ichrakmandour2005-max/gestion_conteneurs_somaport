
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from django.http import HttpResponse



from django.shortcuts import render, redirect, get_object_or_404
from django.contrib import messages
from django.utils import timezone
from .models import Conteneur
from .forms import ConteneurForm
from .services import placer_conteneur, transferer_interne, transferer_externe, AucunEmplacementDisponible

from django.db.models import Count, Q
from django.utils import timezone
from .models import Emplacement

from .services import (
    placer_conteneur, transferer_interne, transferer_externe,
    liberer_emplacements, AucunEmplacementDisponible, ConteneurBloque,
)

def dashboard(request):
    aujourd_hui = timezone.now().date()

    conteneurs_actifs = Conteneur.objects.filter(date_sortie__isnull=True)

    stats = {
        'total_actifs': conteneurs_actifs.count(),
        'entrees_jour': Conteneur.objects.filter(date_entree__date=aujourd_hui).count(),
        'en_transfert': conteneurs_actifs.filter(statut__in=['transfert', 'transfert_externe']).count(),
        'sorties_jour': Conteneur.objects.filter(date_sortie__date=aujourd_hui).count(),
    }

    zones = ['AA', 'AB', 'AC', 'AD']
    occupation_zones = []
    for zone in zones:
        total = Emplacement.objects.filter(zone=zone).count()
        occupees = Emplacement.objects.filter(zone=zone, conteneur__isnull=False).count()
        occupation_zones.append({
            'zone': zone,
            'total': total,
            'occupees': occupees,
            'pourcentage': round((occupees / total) * 100) if total else 0,
            'blocs_pleins': range(min(occupees, 20)),
            'blocs_vides': range(max(total - occupees, 0)) if total <= 20 else range(max(20 - min(occupees, 20), 0)),
        })

    activite_recente = Conteneur.objects.all().order_by('-date_entree')[:8]

    return render(request, 'conteneurs/dashboard.html', {
        'stats': stats,
        'occupation_zones': occupation_zones,
        'activite_recente': activite_recente,
    })

def liste_conteneurs(request):
    q = request.GET.get('q', '').strip().upper()
    conteneurs = Conteneur.objects.all().order_by('-date_entree')

    if q:
        conteneurs = [
            c for c in conteneurs
            if q in c.nom or q in c.code_emplacement()
        ]

    return render(request, 'conteneurs/liste.html', {'conteneurs': conteneurs, 'q': q})


def ajouter_conteneur(request):
    if request.method == 'POST':
        form = ConteneurForm(request.POST)
        if form.is_valid():
            conteneur = form.save(commit=False)
            conteneur.statut = 'entree'
            try:
                conteneur.save()  # déclenche date_entree automatique
                emplacements = placer_conteneur(conteneur)
                codes = ", ".join(e.code for e in emplacements)
                messages.success(
                    request,
                    f"Conteneur {conteneur.nom} enregistré et placé en {codes}."
                )
                return redirect('liste_conteneurs')
            except AucunEmplacementDisponible as e:
                conteneur.delete()  # on annule la création si le placement échoue
                messages.error(request, str(e))
    else:
        form = ConteneurForm()

    return render(request, 'conteneurs/ajouter.html', {'form': form})

def sortir_conteneur(request, conteneur_id):
    conteneur = get_object_or_404(Conteneur, id=conteneur_id)

    if request.method == 'POST':
        type_sortie = request.POST.get('type_sortie')
        try:
            liberer_emplacements(conteneur)
            conteneur.statut = type_sortie
            conteneur.date_sortie = timezone.now()
            conteneur.save()
            messages.success(request, f"Conteneur {conteneur.nom} marqué comme sorti.")
        except ConteneurBloque as e:
            messages.error(request, str(e))
        return redirect('liste_conteneurs')

    return render(request, 'conteneurs/sortir.html', {'conteneur': conteneur})

from .services import placer_conteneur, transferer_interne, transferer_externe, AucunEmplacementDisponible


def transferer_conteneur(request, conteneur_id):
    conteneur = get_object_or_404(Conteneur, id=conteneur_id)

    if request.method == 'POST':
        type_transfert = request.POST.get('type_transfert')

        if type_transfert == 'interne':
            nouvelle_zone = request.POST.get('nouvelle_zone')
            try:
                emplacements = transferer_interne(conteneur, nouvelle_zone)
                codes = ", ".join(e.code for e in emplacements)
                messages.success(request, f"Conteneur {conteneur.nom} transféré en {codes}.")
            except (AucunEmplacementDisponible, ConteneurBloque) as e:
                messages.error(request, str(e))

        elif type_transfert == 'externe':
            try:
                transferer_externe(conteneur)
                messages.success(request, f"Conteneur {conteneur.nom} transféré vers un site externe.")
            except ConteneurBloque as e:
                messages.error(request, str(e))

        return redirect('liste_conteneurs')

    zones = ['AA', 'AB', 'AC', 'AD']
    return render(request, 'conteneurs/transferer.html', {'conteneur': conteneur, 'zones': zones})


def plan_parc(request):
    zones = ['AA', 'AB', 'AC', 'AD']
    rangees = ['A', 'B', 'C', 'D', 'E', 'F', 'G']
    positions = list(range(1, 58, 2))

    emplacements = Emplacement.objects.select_related('conteneur').filter(zone__in=zones)
    index = {(e.zone, e.rangee, e.position): e for e in emplacements}

    zones_data = []
    for zone in zones:
        lignes = []
        for rangee in rangees:
            cellules = []
            skip_next = False
            for position in positions:
                if skip_next:
                    skip_next = False
                    continue

                emp = index.get((zone, rangee, position))

                if emp and emp.conteneur:
                    conteneur = emp.conteneur

                    if conteneur.taille == '40':
                        position_suivante = position + 2
                        emp_suivant = index.get((zone, rangee, position_suivante))
                        if emp_suivant and emp_suivant.conteneur_id == conteneur.id:
                            position_moyenne = (position + position_suivante) // 2
                            cellules.append({
                                'code': f"{zone}{position_moyenne:02d}{rangee}",
                                'occupe': True,
                                'conteneur': conteneur.nom,
                                'etat': conteneur.etat,
                                'large': True,
                            })
                            skip_next = True
                            continue

                    cellules.append({
                        'code': f"{zone}{position:02d}{rangee}",
                        'occupe': True,
                        'conteneur': conteneur.nom,
                        'etat': conteneur.etat,
                        'large': False,
                    })
                else:
                    cellules.append({
                        'code': f"{zone}{position:02d}{rangee}",
                        'occupe': False,
                        'conteneur': None,
                        'etat': None,
                        'large': False,
                    })

            lignes.append({'rangee': rangee, 'cellules': cellules})

        occupees = Emplacement.objects.filter(zone=zone, conteneur__isnull=False).count()
        total = len(rangees) * len(positions)
        zones_data.append({'zone': zone, 'lignes': lignes, 'occupees': occupees, 'total': total})

    return render(request, 'conteneurs/plan.html', {'zones_data': zones_data})

def duree_sejour(request):
    conteneurs = Conteneur.objects.filter(date_sortie__isnull=False).order_by('-date_sortie')
    lignes = []
    total_jours = 0

    for c in conteneurs:
        jours = (c.date_sortie.date() - c.date_entree.date()).days + 1
        lignes.append({'conteneur': c, 'jours': jours})
        total_jours += jours

    moyenne = round(total_jours / len(lignes), 1) if lignes else 0

    return render(request, 'conteneurs/duree_sejour.html', {
        'lignes': lignes,
        'moyenne': moyenne,
    })

def export_duree_sejour(request):
    conteneurs = Conteneur.objects.filter(date_sortie__isnull=False).order_by('-date_sortie')

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Durée de séjour"

    entetes = ["Nom", "Taille", "État", "Date entrée", "Date sortie", "Durée (jours)"]
    ws.append(entetes)

    for cell in ws[1]:
        cell.font = Font(name="Arial", bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="3D3557", end_color="3D3557", fill_type="solid")
        cell.alignment = Alignment(horizontal="center")

    total_jours = 0
    for c in conteneurs:
        jours = (c.date_sortie.date() - c.date_entree.date()).days + 1
        total_jours += jours
        ws.append([
            c.nom,
            c.get_taille_display(),
            c.get_etat_display(),
            c.date_entree.strftime("%d/%m/%Y %H:%M"),
            c.date_sortie.strftime("%d/%m/%Y %H:%M"),
            jours,
        ])

    if conteneurs:
        ws.append([])
        ligne_moyenne = ["", "", "", "", "Moyenne :", round(total_jours / conteneurs.count(), 1)]
        ws.append(ligne_moyenne)
        for cell in ws[ws.max_row]:
            cell.font = Font(name="Arial", bold=True)

    for col in ws.columns:
        longueur_max = max(len(str(cell.value)) if cell.value else 0 for cell in col)
        ws.column_dimensions[col[0].column_letter].width = longueur_max + 4

    for col_lettre in ["A", "B", "C", "D", "E", "F"]:
        for cell in ws[col_lettre]:
            cell.font = Font(name="Arial")

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = 'attachment; filename="duree_sejour.xlsx"'
    wb.save(response)
    return response

from datetime import datetime

def historique(request):
    date_str = request.GET.get('date', '')
    conteneurs = []
    erreur = None

    if date_str:
        try:
            date_selectionnee = datetime.strptime(date_str, '%Y-%m-%d').date()
            conteneurs = Conteneur.objects.filter(
                date_entree__date__lte=date_selectionnee
            ).filter(
                Q(date_sortie__isnull=True) | Q(date_sortie__date__gte=date_selectionnee)
            ).order_by('nom')
        except ValueError:
            erreur = "Date invalide."

    return render(request, 'conteneurs/historique.html', {
        'conteneurs': conteneurs,
        'date_str': date_str,
        'erreur': erreur,
    })